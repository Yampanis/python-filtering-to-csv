import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import undetected_chromedriver as uc
from googleapiclient.discovery import build
from google.oauth2.service_account import Credentials
from oauth2client.service_account import ServiceAccountCredentials
import pandas as pd
import requests
import os
import pickle
from dotenv import load_dotenv
import json
import datetime
import pyperclip
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from selectolax.parser import HTMLParser
from urllib.request import unquote, quote
from urllib.parse import urlparse
import pandas as pd

load_dotenv()
OPEN_API_KEY = os.getenv("OPEN_API_KEY")

def infinite_scroll(driver, max_scrolls=5):
    """Scrolls the page down multiple times to load all content."""
    for _ in range(max_scrolls):
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(2)  

def cleanup_cookies():
    try:
        if os.path.exists("cookies"):
            for file in os.listdir("cookies"):
                if file.endswith(".pkl"):
                    os.remove(os.path.join("cookies", file))
            print("Cleaned up existing cookies")
    except Exception as e:
        print(f"Error cleaning cookies: {e}")

def save_cookies(driver, path):
    """Save cookies to a file"""
    if not os.path.exists('cookies'):
        os.makedirs('cookies')
    with open(path, 'wb') as file:
        pickle.dump(driver.get_cookies(), file)

def load_cookies(driver, path):
    try:
        if os.path.exists(path) and os.path.getsize(path) > 0:
            with open(path, 'rb') as file:
                cookies = pickle.load(file)
                for cookie in cookies:
                    try:
                        driver.add_cookie(cookie)
                    except Exception as e:
                        print(f"Error adding cookie: {e}")
                        continue
            return True
        return False
    except (EOFError, pickle.UnpicklingError) as e:
        print(f"Error loading cookies: {e}")
        try:
            os.remove(path)
            print("Removed corrupted cookie file")
        except OSError:
            pass
        return False

def get_base64_str(source_url):
    try:
        url = urlparse(source_url)
        path = url.path.split("/")
        if url.hostname == "news.google.com" and len(path) > 1 and path[-2] in ["articles", "read"]:
            return {"status": True, "base64_str": path[-1]}
        return {"status": False, "message": "Invalid Google News URL format."}
    except Exception as e:
        return {"status": False, "message": f"Error in get_base64_str: {str(e)}"}

def get_decoding_params(base64_str):
    try:
        url = f"https://news.google.com/articles/{base64_str}"
        response = requests.get(url)
        response.raise_for_status()

        parser = HTMLParser(response.text)
        data_element = parser.css_first("c-wiz > div[jscontroller]")
        if data_element is None:
            return {"status": False, "message": "Failed to fetch data attributes from articles URL."}

        return {
            "status": True,
            "signature": data_element.attributes.get("data-n-a-sg"),
            "timestamp": data_element.attributes.get("data-n-a-ts"),
            "base64_str": base64_str,
        }

    except requests.exceptions.RequestException as req_err:
        try:
            url = f"https://news.google.com/rss/articles/{base64_str}"
            response = requests.get(url)
            response.raise_for_status()

            parser = HTMLParser(response.text)
            data_element = parser.css_first("c-wiz > div[jscontroller]")
            if data_element is None:
                return {"status": False, "message": "Failed to fetch data attributes from RSS URL."}

            return {
                "status": True,
                "signature": data_element.attributes.get("data-n-a-sg"),
                "timestamp": data_element.attributes.get("data-n-a-ts"),
                "base64_str": base64_str,
            }

        except requests.exceptions.RequestException as rss_req_err:
            return {"status": False, "message": f"Request error with RSS URL: {str(rss_req_err)}"}

    except Exception as e:
        return {"status": False, "message": f"Unexpected error in get_decoding_params: {str(e)}"}

def decode_url(signature, timestamp, base64_str):
    try:
        url = "https://news.google.com/_/DotsSplashUi/data/batchexecute"
        payload = [
            "Fbv4je",
            f'["garturlreq",[["X","X",["X","X"],null,null,1,1,"US:en",null,1,null,null,null,null,null,0,1],"X","X",1,[1,1,1],1,1,null,0,0,null,0],"{base64_str}",{timestamp},"{signature}"]',
        ]
        headers = {
            "Content-Type": "application/x-www-form-urlencoded;charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/129.0.0.0 Safari/537.36",
        }

        response = requests.post(
            url,
            headers=headers,
            data=f"f.req={quote(json.dumps([[payload]]))}")
        response.raise_for_status()

        parsed_data = json.loads(response.text.split("\n\n")[1])[:-2]
        decoded_url = json.loads(parsed_data[0][2])[1]

        return {"status": True, "decoded_url": decoded_url}
    except requests.exceptions.RequestException as req_err:
        return {"status": False, "message": f"Request error in decode_url: {str(req_err)}"}
    except (json.JSONDecodeError, IndexError, TypeError) as parse_err:
        return {"status": False, "message": f"Parsing error in decode_url: {str(parse_err)}"}
    except Exception as e:
        return {"status": False, "message": f"Error in decode_url: {str(e)}"}

def decode_google_news_url(source_url, interval=None):
    try:
        base64_response = get_base64_str(source_url)
        if not base64_response["status"]:
            return base64_response

        decoding_params_response = get_decoding_params(base64_response["base64_str"])
        if not decoding_params_response["status"]:
            return decoding_params_response

        if interval:
            time.sleep(interval)

        decoded_url_response = decode_url(
            decoding_params_response["signature"],
            decoding_params_response["timestamp"],
            decoding_params_response["base64_str"],
        )

        return decoded_url_response
    except Exception as e:
        return {"status": False, "message": f"Error in decode_google_news_url: {str(e)}"}

##scopes = [
##    'https://www.googleapis.com/auth/spreadsheets',
##    'https://www.googleapis.com/auth/drive'
##]
##credentials_for_upload_sheet = ServiceAccountCredentials.from_json_keyfile_name(
##    r"C:\Users\the_b\Desktop\Feedly\credentials.json", scopes)
##file = gspread.authorize(credentials_for_upload_sheet)
##sheet3 = file.open("Rory Testing Sheet 2024 ")
##wks3 = sheet3.worksheet("NEW DATA")


try:
    titles_read_df = pd.read_excel(r'titles_to_check.xlsx', sheet_name='Sheet1')
    titles_read = titles_read_df['Titles'].tolist()
except Exception as ex:
    print('Error: ' + str(ex))
    titles_read = []
try:
    negative_titles_df = pd.read_excel(r'negative_titles.xlsx', sheet_name='Sheet1')
    negative_titles_read = negative_titles_df['Titles'].tolist()
except Exception as ex:
    print('Error negative titles: ' + str(ex))
    negative_titles_read = []
try:
    negative_keywords_df = pd.read_excel(r'negatives.xlsx', sheet_name='Sheet1')
    negative_keywords = negative_keywords_df['Negative'].tolist()
except Exception as ex:
    print('Error negatives: '+ str(ex))
    negative_keywords = []
# Convert negative keywords to a set for efficient lookups
negative_keywords_set = set(negative_keywords)

# Define today's date
today_str = datetime.datetime.now().strftime(
    "%a, %d %b %Y %H:%M:%S")  # Format matches article date
if today_str[12:14] == '24':
    today_str = today_str[:12] + '00' + today_str[14:]
new_today_str = datetime.datetime.strptime(today_str, "%a, %d %b %Y %H:%M:%S")
start_range = new_today_str - datetime.timedelta(hours=3)
# Setup Selenium WebDriver
chrome_options = Options()
# chrome_options.add_argument("--start-maximized")
chrome_options.add_argument("--disable-blink-features=AutomationControlled")
chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
chrome_options.add_experimental_option('useAutomationExtension', False)
chrome_options.add_argument("--disable-bluetooth") 
chrome_options.add_argument("--log-level=3")
# Specify the path to chromedriver using Service
chromedriver_path = os.getenv("CHROMEDRIVER_PATH")  # Update with the correct path if necessary
service = Service(executable_path=chromedriver_path)

# Initialize WebDriver with the service and options
driver = webdriver.Chrome(service=service, options=chrome_options)

##options = Options()
##options.add_argument("--start-maximized")
##options.add_argument("--disable-blink-features=AutomationControlled")
##options.add_argument("--ignore-certificate-errors")
##
##driver = uc.Chrome(options=options, driver_executable_path=r'C:\Users\the_b\Desktop\Feedly\browser\chromedriver.exe', version_main=130)

def append_to_excel(existing_file, new_data, sheet_name):
    """Append data to Excel with formatting"""
    try:
        # First read existing data with openpyxl engine
        try:
            existing_data = pd.read_excel(existing_file, sheet_name=None, engine='openpyxl')
        except:
            existing_data = {sheet_name: pd.DataFrame()}
        
        # Combine data
        if sheet_name in existing_data:
            combined_data = pd.concat([existing_data[sheet_name], new_data], ignore_index=True)
        else:
            combined_data = new_data
            
        # Write to Excel using openpyxl engine
        with pd.ExcelWriter(existing_file, engine='openpyxl', mode='w') as writer:
            combined_data.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Get workbook and worksheet objects
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            
            # Add formats
            from openpyxl.styles import Font, PatternFill
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color='D7E4BC', end_color='D7E4BC', fill_type='solid')
            
            for col_num, value in enumerate(combined_data.columns.values):
                # Set column width
                max_length = max(
                    combined_data[value].astype(str).apply(len).max(),
                    len(str(value))
                ) + 2
                worksheet.column_dimensions[get_column_letter(col_num + 1)].width = max_length
                
                # Format header using new style
                cell = worksheet.cell(row=1, column=col_num + 1)
                cell.font = header_font
                cell.fill = header_fill
                
    except Exception as e:
        print(f"Error in append_to_excel: {e}")

def adjust_column_width(workbook_path):
    """Adjust column widths based on content"""
    try:
        # Load the workbook
        workbook = load_workbook(workbook_path)
        worksheet = workbook.active
        
        # Iterate through columns
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            # Find longest content in column
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Set width with some padding
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save the workbook
        workbook.save(workbook_path)
        print(f"Adjusted column widths in {workbook_path}")
    except Exception as e:
        print(f"Error adjusting column widths: {e}")

# Function to check if a title contains any negative keywords
def contains_negative_keywords(title, keywords_set):
    # Split title into words and check intersection with negative keywords
    words_in_title = set(title.lower().split())
    return not words_in_title.isdisjoint(keywords_set)

def check_title_against_keywords(title, negative_keywords):
    """
    Check if a title contains any keyword (with word boundaries).
    """
    title_lower = title.lower()
    for keyword in negative_keywords:
        # Use word boundary (\b) regex for word matching
        keyword_pattern = r'\b' + re.escape(keyword.lower()) + r'\b'
        if re.search(keyword_pattern, title_lower):
            print(f"Title: {title_lower} contains keyword: {keyword_pattern}")
            return False
    return True

def url_contains_keyword(url, negative_keywords):
    """
    Check if a URL contains any keyword as a substring.
    Strips spaces and ensures proper matches.
    """
    url_lower = url.strip().lower()  # Normalize URL
    for keyword in negative_keywords:
        # Normalize keyword (strip spaces, lowercase)
        keyword_clean = keyword.lower()
        # Substring matching
        if keyword_clean in url_lower:
            print(f"Url: {url_lower} contains keyword: {keyword_clean}")
            return False
    return True


def feedly_login(driver, email, password):
    cookies_path = "cookies/feedly_cookies.pkl"
    driver.get("https://feedly.com/i/discover")
    driver.set_page_load_timeout(30)  
    
    time.sleep(3.5)

    cookie_login_successful = False
    try: 
        if load_cookies(driver, cookies_path):
            driver.refresh()
            try:
                WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located((By.ID, "feedlyFrame"))
                )
                cookie_login_successful = True
                return True
            except Exception as e:
                print(f"Cookie login failed: {e}")
    except Exception as e:
        print(f"Error loading cookies: {e}")

    if not cookie_login_successful:    
        try:
            print ("Logging in with Gmail and password")
            login_button = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "(//button[contains(.,'Log In')])[1]"))
            )
            login_button.click()
            time.sleep(2)

            google_login = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//a[contains(.,'with Google')]")
            ))
            google_login.click()

            email_input = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//input[@type='email']"))
            )
            email_input.send_keys(email)
            email_input.send_keys(Keys.ENTER)
            time.sleep(2)

            password_input = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//input[@type='password']"))
            )
            password_input.send_keys(password)
            password_input.send_keys(Keys.ENTER)
            time.sleep(10)
            print("Logged in successfully!!!")
            # save_cookies(driver, cookies_path)
        except Exception as e:
            print(f"Error in feedly_login: {str(e)}")
            driver.quit()

def login_to_chatgpt_com(driver):
    try:
        login_button = WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.XPATH, '//button[@data-testid="login-button"]'))
        )
        login_button.click()
        time.sleep(5)
    except:
        print("Login button not found on chatgpt.com.")
        return
    # Attemp to login
    try:
        email = WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.XPATH, '//input[@id="email-input"]'))
        )#driver.find_element(By.XPATH, "//input[@id='email-input']")
        email.click()
        email.send_keys("rory@thebestreputation.com")
        email.send_keys(Keys.RETURN)
        time.sleep(5)
    except Exception as e:
        print("Unable to process username! Error: " + str(e))
        return
    # Attemp password
    try:
        passwd = WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.XPATH, '//input[@id="password"]'))
        )
        passwd.click()
        passwd.send_keys("Barbeque2045!")
        passwd.send_keys(Keys.RETURN)
        time.sleep(5)
    except Exception as e:
        print("Unable to process password to login! Error: " + str(e))
        return

def ask_chatgpt_com(brw, prompt):
    brw.get("https://chatgpt.com/")
    time.sleep(5)
    try:
        modal_cls = brw.find_element(By.XPATH, "//*[contains(@id,'radix-:r')]/div/div/a")
        modal_cls.close()
        time.sleep(2)
    except Exception as ex:
        pass
##    WebDriverWait(brw, 10).until(EC.presence_of_element_located((By.ID, 'prompt-textarea')))

    # Locate the input box and submit the prompt
    # Use JavaScript to reveal the hidden input box if necessary
    try:
        input_box = WebDriverWait(brw, 10).until(
            EC.element_to_be_clickable((By.ID, 'prompt-textarea'))
        )
        brw.execute_script("arguments[0].style.display = 'block';", input_box)  # Force display
        time.sleep(3)  # Allow JavaScript change to take effect
        input_box.clear()
        pyperclip.copy(prompt)

        # Click to focus the input box if necessary
        input_box.click()
        
        input_box.send_keys(Keys.CONTROL, 'v')
##        time.sleep(45)  # Wait for ChatGPT to generate a response
        input_box.send_keys(Keys.RETURN)
    except Exception as e:
        return f"Error: Unable to locate or interact with the input box. Details: {e}"

    # Wait until response is generated
    try:
        time.sleep(85)
    except Exception as e:
        return f"Error: Response generation timed out. Details: {e}"

    # Capture and return the response text
    try:
        response_elements = brw.find_elements(By.CSS_SELECTOR, ".markdown div.border-token-border-medium div.overflow-y-auto")
        if len(response_elements) > 1:
            response_text = response_elements[-1].get_attribute('outerText') if response_elements else "No response found"
        else:
            response_text = response_elements[0].get_attribute('outerText') if response_elements else "No response found"
    except Exception as ex:
        try:
            response_elements = brw.find_elements(By.CSS_SELECTOR, ".markdown p")
            resp_elems = ''
            if len(response_elements) > 1:
                for resel in resp_elems:
                    resp_elems += resel.get_attribute('outerText') + '\n'
                resp_elems = resp_elems.strip()
                response_text = resp_elems
            else:
                response_text = response_elements[0].get_attribute('outerText')
        except Exception as ex:
            print('Response text is empty!')
            response_text = ''
    if '#' in response_text:
        lines = response_text.strip().split('\n')
        # Initialize a list to hold the articles
        articles = []
        for line in lines:
            parts = line.split('#')  # Split by tab character
            if len(parts) >= 4:
                try:
                    try:
                        url = re.sub(r'\d. ','',parts[0].strip())
                    except:
                        url = parts[0].strip()
                except Exception as e:
                    url = 'N/A'
                try:
                    title = parts[1].strip()
                except Exception as e:
                    title = 'N/A'
                try:
                    description = parts[2].strip()
                except Exception as e:
                    description = 'N/A'
                try:
                    reach_out = parts[3].strip()
                except Exception as e:
                    reach_out = 'N/A'
                try:
                    reasons = parts[4].strip()
                except Exception as e:
                    reasons = 'N/A'
                try:
                    keywords = parts[5].strip()  # All keywords
                except Exception as e:
                    keywords = 'N/A'
                try:
                    if '|' in parts[6]:
                        location = re.sub(r'\|',', ', parts[6])
                    else:
                        location = parts[6].strip()
                except Exception as e:
                    location = 'N/A'

                # Create a dictionary for the article
                article = {
                    "url": url,
                    "title": title,
                    "description": description,
                    "reach_out": reach_out,
                    "reasons": reasons,
                    "keywords": keywords,
                    "location": location,
                }

                # Add the article dictionary to the list
                articles.append(article)
            # Convert the list of articles to JSON format
        json_output = json.dumps(articles, indent=4)
        return json_output
    elif '\t' in response_text:
        lines = response_text.strip().split('\n')
        # Initialize a list to hold the articles
        articles = []
        for line in lines:
            parts = line.split('\t')  # Split by tab character
            if len(parts) >= 4:
                try:
                    try:
                        url = re.sub(r'\d. ','',parts[0].strip())
                    except:
                        url = parts[0].strip()
                except Exception as e:
                    url = 'N/A'
                try:
                    title = parts[1].strip()
                except Exception as e:
                    title = 'N/A'
                try:
                    description = parts[2].strip()
                except Exception as e:
                    description = 'N/A'
                try:
                    reach_out = parts[3].strip()
                except Exception as e:
                    reach_out = 'N/A'
                try:
                    reasons = parts[4].strip()
                except Exception as e:
                    reasons = 'N/A'
                try:
                    keywords = parts[5].strip()  # All keywords
                except Exception as e:
                    keywords = 'N/A'
                try:
                    if '|' in parts[6]:
                        location = re.sub(r'\|',', ', parts[6])
                    else:
                        location = parts[6].strip()
                except Exception as e:
                    location = 'N/A'

                # Create a dictionary for the article
                article = {
                    "url": url,
                    "title": title,
                    "description": description,
                    "reach_out": reach_out,
                    "reasons": reasons,
                    "keywords": keywords,
                    "location": location,
                }

                # Add the article dictionary to the list
                articles.append(article)
        # Convert the list of articles to JSON format
        json_output = json.dumps(articles, indent=4)
        return json_output
    else:
        print("Unexpected response structure:", response_text)
        return None
##    except Exception as e:
##        return f"Error: Unable to retrieve response. Details: {e}"

def call_chatgpt_api(prompt, api_key):
    url = "https://api.openai.com/v1/chat/completions"
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json"
    }
    # Dynamically set max_tokens to stay within model limits
    prompt_tokens = len(prompt.split())
    max_tokens_response = min(4096 - prompt_tokens, 3700)
    data = {
        "model": "gpt-4",  # or use the model of your choice
        "messages": [{
            "role": "user",
            "content": prompt
        }],
        "max_tokens": 100  # adjust the number of tokens as needed
    }
    new_data = ''

    max_retries = 5
    for attempt in range(max_retries):
        if new_data == '':
            response = requests.post(url, headers=headers, json=data)
        else:
            response = requests.post(url, headers=headers, json=new_data)
        time.sleep(7)

        print(f"Attempt {attempt + 1}: Status Code: {response.status_code}"
              )  # Log the status code

        if response.status_code == 429:  # Rate limit exceeded
            wait_time = 2**attempt  # Exponential backoff
            print(f"Rate limit exceeded. Waiting for {wait_time} seconds...")
            time.sleep(wait_time)
            continue  # Retry the request
        elif response.status_code == 200:  # Successful response
            try:
                response_json = response.json()
                # Ensure the expected structure is present
                if 'choices' in response_json and len(
                        response_json['choices']) > 0:
                    content = response_json['choices'][0]['message']['content']
                    if not content.strip():  # Check if content is empty
                        print("Received empty response, retrying...")
                        continue  # Retry if the response is empty
                    # Split the raw data into individual lines
                    lines = content.strip().split('\n')
                    # Initialize a list to hold the articles
                    articles = []

                    # Process each line to extract data
                    for line in lines:
                        parts = line.split('#')  # Split by tab character
                        if len(parts) >= 4:
                            try:
                                try:
                                    url = re.sub(r'\d. ','',parts[0].strip())
                                except:
                                    url = parts[0].strip()
                            except Exception as e:
                                url = '-'
                            try:
                                title = parts[1].strip()
                            except Exception as e:
                                title = '-'
                            try:
                                description = parts[2].strip()
                            except Exception as e:
                                description = '-'
                            try:
                                reach_out = parts[3].strip()
                            except Exception as e:
                                reach_out = '-'
                            try:
                                reasons = parts[4].strip()
                            except Exception as e:
                                reasons = '-'
                            try:
                                keywords = parts[5].strip()  # All keywords
                            except Exception as e:
                                keywords = '-'
                            try:
                                location = parts[6].strip()
                            except Exception as e:
                                location = '-'

                            # Create a dictionary for the article
                            article = {
                                "url": url,
                                "title": title,
                                "description": description,
                                "reach_out": reach_out,
                                "reasons": reasons,
                                "keywords": keywords,
                                "location": location,
                            }

                            # Add the article dictionary to the list
                            articles.append(article)
                    # Convert the list of articles to JSON format
                    json_output = json.dumps(articles, indent=4)
                    return json_output
                else:
                    print("Unexpected response structure:", response_json)
                    return None
            except json.JSONDecodeError:
                print("Failed to decode JSON response:", response.text)
                return None
        else:
            try:
                new_data = {
                    "model": "gpt-3.5-turbo",  # or use the model of your choice
                    "messages": [{
                        "role": "user",
                        "content": prompt
                    }],
                    "max_tokens": max_tokens_response  # adjust the number of tokens as needed
                }
                continue
            except Exception as e:
                print(
                    f"Unexpected status code: {response.status_code}, response: {response.text}"
                )
                response.raise_for_status(
                )  # Raise an error for other bad responses


def scroll_down(driver, element_selector):
    try:
        # Find the element (e.g., <main> tag or another container) and scroll within it
        main_element = WebDriverWait(driver, 5).until(
            EC.visibility_of_element_located((By.XPATH, element_selector))
        )
        # Scroll to the bottom of the element
        driver.execute_script(
            "arguments[0].scrollTo(0, arguments[0].scrollHeight);", main_element)
        time.sleep(1.5)
    except Exception as e:
        pass


def scrape_today_articles(driver):
    new_articles = []
    batch_size = 500
    last_height = driver.execute_script(
        "return document.querySelector('#feedlyFrame').scrollHeight")
    '''while True:
        # Scroll down and wait for page load
        scroll_down(driver, "//*[@id='feedlyFrame']")

        # Check if we reached the bottom
        new_height = driver.execute_script(
            "return document.querySelector('#feedlyFrame').scrollHeight")
        if new_height == last_height:
            break
        last_height = new_height
        time.sleep(1.5)'''

    try:
        articles = driver.find_elements(By.CLASS_NAME, "entry.magazine")
        time.sleep(3)
        for article in articles:
            try:
                date_span = article.find_element(By.CSS_SELECTOR, "span[title*='Published']")
                article_date = date_span.get_attribute("title")
                new_article_date = re.sub(r'\n', '', str(article_date)).strip()
                new_article_date = re.sub(r'.*Received: | GMT.*', '', new_article_date).strip()
                if new_article_date[17:19] == '24':
                    new_article_date = new_article_date[:17] + '00' + new_article_date[19:]
                new_article_date_conv = datetime.datetime.strptime(new_article_date, "%a, %d %b %Y %H:%M:%S")
                if start_range <= new_article_date_conv <= new_today_str:
                    title = article.find_element(By.CSS_SELECTOR, "a.EntryTitleLink").text
                    link = article.find_element(By.CSS_SELECTOR, "a.EntryTitleLink").get_attribute("href")
                    
                    if len(new_articles) % batch_size == 0:
                        driver.execute_script("window.localStorage.clear();")
                        driver.execute_script("window.sessionStorage.clear();")

                    article = (title, link)
                    if article not in new_articles:
                        new_articles.append(article)
                        
            except Exception as e:
                print('Error for article: ' + str(e))
        time.sleep(1.5)
        driver.get(
            'https://feedly.com/i/collection/content/user/9e62dc2d-90e6-453b-88f4-47b630b9a4aa/category/global.all'
            #'https://feedly.com/i/collection/content/user/9fa377e1-a6c0-4f6a-8e98-ab3cc30fd0c3/category/global.all' #m78077439@gmail.com
            #'https://feedly.com/i/collection/content/user/9e62dc2d-90e6-453b-88f4-47b630b9a4aa/category/global.all' #m08067064@gmail.com
            
        )
        time.sleep(10)
        counter = 0
        new_last_height = driver.execute_script(
            "return document.querySelector('#feedlyFrame').scrollHeight")
        while counter < 30:
            # Scroll down and wait for page load
            scroll_down(driver, "//*[@id='feedlyFrame']")
            # Check if we reached the bottom
            new_height1 = driver.execute_script( 
                "return document.querySelector('#feedlyFrame').scrollHeight")
            if new_height1 == new_last_height:
                break
            new_last_height = new_height1
            counter += 1
            time.sleep(1.5)

        new_articles1 = driver.find_elements(By.CLASS_NAME, "entry.magazine")
        for article in new_articles1:
            try:
                date_span1 = article.find_element(By.CSS_SELECTOR, "span[title*='Published']")
                article_date1 = date_span1.get_attribute("title")
                new_article_date1 = re.sub(r'\n', '', str(article_date1)).strip()
                new_article_date1 = re.sub(r'.*Received: | GMT.*', '', new_article_date1).strip()
                if new_article_date1[17:19] == '24':
                    new_article_date1 = new_article_date1[:17] + '00' + new_article_date1[19:]
                new_article_date1_conv = datetime.datetime.strptime(new_article_date1, "%a, %d %b %Y %H:%M:%S")
                #if 1 == 1:
                if start_range <= new_article_date1_conv <= new_today_str:
                    title = article.find_element(By.CSS_SELECTOR, "a.EntryTitleLink").text
                    link = article.find_element(By.CSS_SELECTOR, "a.EntryTitleLink").get_attribute("href")
                    
                    article = (title, link)
                    if article not in new_articles:
                        new_articles.append(article)
            except Exception as e:
                print('Error for article: ' + str(e))
        
        print("Debugger!!!")
    except Exception as e:
        print('Error execution: ' + str(e))
        pass
    finally:
        driver.quit()

    return new_articles

def process_articles_batch(unique_new_articles, batch_size=50):
    decoded_articles = []
    titles = []
    titles_neg = []
    pg_links = []
    existing_titles = set()

    for i in range(0, len(unique_new_articles), batch_size):
        batch = unique_new_articles[i:i + batch_size]

        for article in batch:
            title, url = str(article[0]), str(article[1])
            if title in existing_titles:
                continue
            existing_titles.add(title)

            if 'news.google' in url:
                try:
                    decoded_url = decode_google_news_url(url, interval=0.1)

                    if decoded_url.get("status"):
                        final_url = decoded_url["decoded_url"]
                        print("Decoded url: ", final_url)
                        decoded_articles.append((title, final_url))
                        
                        # Batch keyword checking
                        if (url_contains_keyword(final_url, negative_keywords) or 
                            check_title_against_keywords(title, negative_keywords)):
                            titles_neg.append(title)
                        else:
                            pg_links.append(final_url)
                            titles.append(title)
                    else:
                        decoded_articles.append((title, url))
                        pg_links.append(url)
                        titles.append(title)
                    
                except Exception as e:
                    print('Error trying to convert google news link: ' + str(e))
                    pg_links.append(url)
                    titles.append(title)
            else:
                pg_links.append(url)
                titles.append(title)
            
    return {
        'decoded_articles': decoded_articles,
        'titles': titles,
        'pg_links': pg_links,
        'titles_neg': titles_neg
    }

def main(email, password):
    df_no_duplicates = pd.DataFrame()

    # Define headers and empty data
    if not os.path.exists(r"Rory Testing Sheet 2024.xlsx"):
        headers = ["Url","Title","Description","Reach Out","Reasons", "Keywords","Location","NOTES"]
        data = []
        df = pd.DataFrame(data, columns=headers)
        # Save to an Excel file
        df.to_excel(r"Rory Testing Sheet 2024.xlsx", index=False, engine="openpyxl")
        
        adjust_column_width(r"Rory Testing Sheet 2024.xlsx")
    
    titles_to_check = []
    existing_titles = []
    titles = []
    pg_links = []
    pg_links_check = []
    titles_check = []
    titles_neg = []
    try:
        feedly_login(driver, email, password)
        time.sleep(1.4) 
        articles = scrape_today_articles(driver)

        # Process articles
        if articles:
            print('Total collected articles: ' + str(len(articles)))
            unique_new_articles_neg = [article for article in articles if article[0] not in titles_read]
            print('Total unique articles before negatives check: ' + str(len(unique_new_articles_neg)))
            unique_new_articles = [article for article in unique_new_articles_neg if article[0] not in negative_titles_read]
            print('Total unique articles after negatives check: ' + str(len(unique_new_articles)))
            total_titles = len(titles_read)
            print('Total titles read: ' + str(total_titles))

            titles_read_set = set(titles_read)
            negative_titles_set = set(negative_titles_read)

            # Filter articles in one pass
            unique_new_articles = [
                article for article in articles 
                if article[0] not in titles_read_set and article[0] not in negative_titles_set
            ]

            results = process_articles_batch(unique_new_articles)

            # Write to CSV immediately after processing
            try:
                articles_df = pd.DataFrame(results['decoded_articles'], columns=['Title', 'URL'])
                articles_df.to_csv(r'feedly_articles.csv', mode='a', index=False)
                print(f"Successfully wrote {len(results['decoded_articles'])} articles")
                
                # Create DataFrames for titles
                if results['titles']:
                    df_titles = pd.DataFrame({"Titles": results['titles']})
                    append_to_excel(r'titles_to_check.xlsx', df_titles, 'Sheet1')
                    
                if results['titles_neg']:
                    df_titles_neg = pd.DataFrame({"Titles": results['titles_neg']})
                    append_to_excel(r'negative_titles.xlsx', df_titles_neg, 'Sheet1')

                excel_data = pd.DataFrame({
                    "URL": [article[1] for article in results['decoded_articles']],
                    "Title": [article[0] for article in results['decoded_articles']]
                })

                append_to_excel(r"Rory Testing Sheet 2024.xlsx", excel_data, 'Sheet1')
                print(f"Successfully wrote {len(results['decoded_articles'])} articles to Excel")
                    
            except Exception as e:
                print(f"Error writing to files: {str(e)}")

    finally:
        driver.quit()
        driver.execute_script("window.localStorage.clear();")
        driver.execute_script("window.sessionStorage.clear();")
        driver.execute_script("console.clear();")

if __name__ == "__main__":
    start = time.time()
    email = "m08067064@gmail.com"
    password = "thebestrep2025"
    main(email, password)
    end = time.time()
    print("Time Taken: {:.6f}s".format(end - start))
