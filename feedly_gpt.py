import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
import requests
import os
import pickle
from dotenv import load_dotenv
import json
import datetime
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from selectolax.parser import HTMLParser
from urllib.request import quote
from urllib.parse import urlparse
import pandas as pd
import logging
from openai import OpenAI

from concurrent.futures import ThreadPoolExecutor, as_completed
import threading
from queue import Queue

load_dotenv()
OPEN_API_KEY = os.getenv("OPEN_API_KEY")

logging.basicConfig(
    filename='feedly.log',                 # File to write logs to
    level=logging.INFO,                 # Minimum level of logs to record
    format='%(asctime)s - %(levelname)s - %(message)s',  # Log message format
    datefmt='%Y-%m-%d %H:%M:%S'          # Date/time format
)

titles_read = []
negative_titles_read = []
negative_keywords = []

def infinite_scroll(driver, max_scrolls=5):
    """Scrolls the page down multiple times to load all content."""
    for _ in range(max_scrolls):
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(2)  

def cleanup_cookies():
    try:
        if os.path.exists("cookies"):
            for file in os.listdir("cookies"):
                if file.endswith(".pkl"):
                    os.remove(os.path.join("cookies", file))
            print("Cleaned up existing cookies")
    except Exception as e:
        print(f"Error cleaning cookies: {e}")

def save_cookies(driver, path):
    """Save cookies to a file"""
    if not os.path.exists('cookies'):
        os.makedirs('cookies')
    with open(path, 'wb') as file:
        pickle.dump(driver.get_cookies(), file)

def load_cookies(driver, path):
    try:
        if os.path.exists(path) and os.path.getsize(path) > 0:
            with open(path, 'rb') as file:
                cookies = pickle.load(file)
                for cookie in cookies:
                    try:
                        driver.add_cookie(cookie)
                    except Exception as e:
                        print(f"Error adding cookie: {e}")
                        continue
            return True
        return False
    except (EOFError, pickle.UnpicklingError) as e:
        print(f"Error loading cookies: {e}")
        try:
            os.remove(path)
            print("Removed corrupted cookie file")
        except OSError:
            pass
        return False

def get_base64_str(source_url):
    try:
        url = urlparse(source_url)
        path = url.path.split("/")
        if url.hostname == "news.google.com" and len(path) > 1 and path[-2] in ["articles", "read"]:
            return {"status": True, "base64_str": path[-1]}
        return {"status": False, "message": "Invalid Google News URL format."}
    except Exception as e:
        return {"status": False, "message": f"Error in get_base64_str: {str(e)}"}

def get_decoding_params(base64_str):
    try:
        url = f"https://news.google.com/articles/{base64_str}"
        response = requests.get(url)
        response.raise_for_status()

        parser = HTMLParser(response.text)
        data_element = parser.css_first("c-wiz > div[jscontroller]")
        if data_element is None:
            return {"status": False, "message": "Failed to fetch data attributes from articles URL."}

        return {
            "status": True,
            "signature": data_element.attributes.get("data-n-a-sg"),
            "timestamp": data_element.attributes.get("data-n-a-ts"),
            "base64_str": base64_str,
        }

    except requests.exceptions.RequestException as req_err:
        try:
            url = f"https://news.google.com/rss/articles/{base64_str}"
            response = requests.get(url)
            response.raise_for_status()

            parser = HTMLParser(response.text)
            data_element = parser.css_first("c-wiz > div[jscontroller]")
            if data_element is None:
                return {"status": False, "message": "Failed to fetch data attributes from RSS URL."}

            return {
                "status": True,
                "signature": data_element.attributes.get("data-n-a-sg"),
                "timestamp": data_element.attributes.get("data-n-a-ts"),
                "base64_str": base64_str,
            }

        except requests.exceptions.RequestException as rss_req_err:
            return {"status": False, "message": f"Request error with RSS URL: {str(rss_req_err)}"}

    except Exception as e:
        return {"status": False, "message": f"Unexpected error in get_decoding_params: {str(e)}"}

def decode_url(signature, timestamp, base64_str):
    try:
        url = "https://news.google.com/_/DotsSplashUi/data/batchexecute"
        payload = [
            "Fbv4je",
            f'["garturlreq",[["X","X",["X","X"],null,null,1,1,"US:en",null,1,null,null,null,null,null,0,1],"X","X",1,[1,1,1],1,1,null,0,0,null,0],"{base64_str}",{timestamp},"{signature}"]',
        ]
        headers = {
            "Content-Type": "application/x-www-form-urlencoded;charset=UTF-8",
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/129.0.0.0 Safari/537.36",
        }

        response = requests.post(
            url,
            headers=headers,
            data=f"f.req={quote(json.dumps([[payload]]))}")
        response.raise_for_status()

        parsed_data = json.loads(response.text.split("\n\n")[1])[:-2]
        decoded_url = json.loads(parsed_data[0][2])[1]

        return {"status": True, "decoded_url": decoded_url}
    except requests.exceptions.RequestException as req_err:
        return {"status": False, "message": f"Request error in decode_url: {str(req_err)}"}
    except (json.JSONDecodeError, IndexError, TypeError) as parse_err:
        return {"status": False, "message": f"Parsing error in decode_url: {str(parse_err)}"}
    except Exception as e:
        return {"status": False, "message": f"Error in decode_url: {str(e)}"}

def decode_google_news_url(source_url, interval=None):
    try:
        base64_response = get_base64_str(source_url)
        if not base64_response["status"]:
            return base64_response

        decoding_params_response = get_decoding_params(base64_response["base64_str"])
        if not decoding_params_response["status"]:
            return decoding_params_response

        if interval:
            time.sleep(interval)

        decoded_url_response = decode_url(
            decoding_params_response["signature"],
            decoding_params_response["timestamp"],
            decoding_params_response["base64_str"],
        )

        return decoded_url_response
    except Exception as e:
        return {"status": False, "message": f"Error in decode_google_news_url: {str(e)}"}

def initialize_global_variables():
    global titles_read, negative_titles_read, negative_keywords
    try:
        titles_read_df = pd.read_excel(r'titles_to_check.xlsx', sheet_name='Sheet1')
        titles_read = titles_read_df['Titles'].tolist()
    except Exception as ex:
        print('Error: ' + str(ex))
        titles_read = []
    try:
        negative_titles_df = pd.read_excel(r'negative_titles.xlsx', sheet_name='Sheet1')
        negative_titles_read = negative_titles_df['Titles'].tolist()
    except Exception as ex:
        print('Error negative titles: ' + str(ex))
        negative_titles_read = []
    try:
        negative_keywords_df = pd.read_excel(r'negatives.xlsx', sheet_name='Sheet1')
        negative_keywords = negative_keywords_df['Negative'].tolist()
        logging.info(f"'{len(negative_keywords)}' negative keywords")
    except Exception as ex:
        print('Error negatives: '+ str(ex))
        negative_keywords = []
    
    return titles_read, negative_titles_read, negative_keywords

# Define today's date
today_str = datetime.datetime.now().strftime(
    "%a, %d %b %Y %H:%M:%S")  # Format matches article date
if today_str[12:14] == '24':
    today_str = today_str[:12] + '00' + today_str[14:]
new_today_str = datetime.datetime.strptime(today_str, "%a, %d %b %Y %H:%M:%S")

logging.info(f"'{new_today_str}' new_today_str")
print(new_today_str)
start_range = new_today_str - datetime.timedelta(hours=3)
# Setup Selenium WebDriver
chrome_options = Options()
# chrome_options.add_argument("--start-maximized")
chrome_options.add_argument("--disable-blink-features=AutomationControlled")
chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
chrome_options.add_experimental_option('useAutomationExtension', False)
chrome_options.add_argument("--disable-bluetooth") 
chrome_options.add_argument("--log-level=3")
# Specify the path to chromedriver using Service
chromedriver_path = os.getenv("CHROMEDRIVER_PATH")  # Update with the correct path if necessary
service = Service(executable_path=chromedriver_path)

# Initialize WebDriver with the service and options
driver = webdriver.Chrome(service=service, options=chrome_options)

def create_gpt_prompt(article_batch):
    """Create structured prompt for GPT analysis"""
    prompt = """Analyze each article and provide information in exactly this format:
    URL#Title#Description#ReachOut#Reasons#Keywords#Location

    Important: Do not use '#' character anywhere in the content, use commas or semicolons instead.
    
    For each article provide:
    - Description: 100-200 word summary (no '#' characters)
    - ReachOut: Key person or organization to contact
    - Reasons: Why this is relevant (use commas instead of '#')
    - Keywords: Up to 20 important terms (comma separated)
    - Location: Geographic location mentioned
    
    Format each response on a new line with exactly 6 '#' separators.
    """
    
    for title, url in article_batch:
        prompt += f"\nArticle: {url}\nTitle: {title}\n"
    
    return prompt

def call_gpt_api(prompt):
    """Call GPT API using GPT-4 Turbo model for faster processing"""
    try:
        client = OpenAI(api_key=OPEN_API_KEY)
        response = client.chat.completions.create(
            model="gpt-4-1106-preview",  # Changed to GPT-4 Turbo
            messages=[{
                "role": "user",
                "content": prompt
            }],
            max_tokens=1000,
            temperature=0.7,
            response_format={ "type": "text" }  # Force text response for faster processing
        )
        
        if response.choices:
            content = response.choices[0].message.content
            return parse_gpt_api_response_content(content)
        else:
            print("No response content from GPT")
            return None
            
    except Exception as e:
        print(f"OpenAI API call error: {str(e)}")
        return None

def parse_gpt_api_response_content(content):
    """Parse GPT response content directly"""
    try:
        articles = []
        lines = content.strip().split('\n')

        for line in lines:
            if '#' not in line:
                continue

            parts = line.split('#')
            if len(parts) >= 7:
                article = {
                    "url": parts[0].strip(),
                    "title": parts[1].strip(),
                    "description": parts[2].strip(),
                    "reach_out": parts[3].strip(),
                    "reasons": parts[4].strip(),
                    "keywords": parts[5].strip(),
                    "location": parts[6].strip()
                }
                articles.append(article)

        return json.dumps(articles, indent=4) if articles else None

    except Exception as e:
        print(f"Response parsing error: {str(e)}")
        return None

def append_to_excel(existing_file, new_data, sheet_name):
    """Append data to Excel with optimized column widths"""
    try:
        # First read existing data
        try:
            existing_data = pd.read_excel(existing_file, sheet_name=None, engine='openpyxl')
        except:
            existing_data = {sheet_name: pd.DataFrame()}
        
        # Combine data
        if sheet_name in existing_data:
            combined_data = pd.concat([existing_data[sheet_name], new_data], ignore_index=True)
        else:
            combined_data = new_data
            
        # Write to Excel
        with pd.ExcelWriter(existing_file, engine='openpyxl', mode='w') as writer:
            combined_data.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Get workbook and worksheet objects
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            
            # Add formats
            from openpyxl.styles import Font, PatternFill
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color='D7E4BC', end_color='D7E4BC', fill_type='solid')
            
            # Updated column widths with more reasonable values
            column_widths = {
                'URL': 25,           # Reduced from 30
                'Title': 40,         # Reduced from 60
                'Description': 30,    # Reduced from 40
                'Reach Out': 12,     # Reduced from 15
                'Reasons': 20,       # Reduced from 25
                'Keywords': 12,      # Reduced from 15
                'Location': 12,      # Reduced from 15
                'NOTES': 15          # Reduced from 25
            }
            
            # Apply column widths and formatting
            for col_num, value in enumerate(combined_data.columns.values):
                column_letter = get_column_letter(col_num + 1)
                
                # Set fixed column width based on predefined values
                if value in column_widths:
                    worksheet.column_dimensions[column_letter].width = column_widths[value]
                else:
                    # For any other columns, set a reasonable default
                    worksheet.column_dimensions[column_letter].width = 15
                
                # Format header
                cell = worksheet.cell(row=1, column=col_num + 1)
                cell.font = header_font
                cell.fill = header_fill
                
    except Exception as e:
        print(f"Error in append_to_excel: {e}")

def adjust_column_width(workbook_path):
    """Adjust column widths based on content"""
    try:
        # Load the workbook
        workbook = load_workbook(workbook_path)
        worksheet = workbook.active
        
        # Iterate through columns
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            # Find longest content in column
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Set width with some padding
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save the workbook
        workbook.save(workbook_path)
        print(f"Adjusted column widths in {workbook_path}")
    except Exception as e:
        print(f"Error adjusting column widths: {e}")

def is_check_title_against_keywords(title, negative_keywords):
    title_lower = title.lower().strip()
    
    for keyword in negative_keywords:
        keyword_lower = keyword.lower().strip()
        
        keyword_pattern = r'(?:^|[\s,.\-!?()\[\]{}"]|\b)' + re.escape(keyword_lower)
        
        if re.search(keyword_pattern, title_lower):
            print(f"Found negative keyword '{keyword_lower}' in title: {title}")
            return True
            
    return False

def is_url_contains_keyword(url, negative_keywords):
    url_lower = url.strip().lower()
    url_parts = re.split(r'[/\-_.]', url_lower)
    
    for keyword in negative_keywords:
        keyword_lower = keyword.lower().strip()
        
        keyword_pattern = r'(?:^|[/\-_.]|\b)' + re.escape(keyword_lower)
        
        if re.search(keyword_pattern, url_lower):
            print(f"Found negative keyword '{keyword_lower}' in URL: {url}")
            return True
        
        for part in url_parts:
            if part and re.search(f'^{re.escape(keyword_lower)}', part):
                print(f"Found negative keyword '{keyword_lower}' in URL part: {part}")
                return True
                
    return False


def feedly_login(driver, email, password):
    """Login to Feedly with proper window handling"""
    cookies_path = "cookies/feedly_cookies.pkl"
    driver.get("https://feedly.com/i/discover")
    driver.set_page_load_timeout(30)  
    time.sleep(5)

    cookie_login_successful = False
    try: 
        if load_cookies(driver, cookies_path):
            driver.refresh()
            try:
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.ID, "feedlyFrame"))
                )
                cookie_login_successful = True
                return True
            except Exception as e:
                print(f"Cookie login failed: {e}")
    except Exception as e:
        print(f"Error loading cookies: {e}")

    if not cookie_login_successful:    
        try:
            print ("Logging in with Gmail and password")
            login_button = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "(//button[contains(.,'Log In')])[1]"))
            )
            login_button.click()
            time.sleep(2)

            google_login = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//a[contains(.,'with Google')]")
            ))
            google_login.click()
            time.sleep(2)

            # Handle Google login popup
            windows = driver.window_handles
            if len(windows) > 1:
                driver.switch_to.window(windows[-1])

            # Enter email
            email_input = WebDriverWait(driver, 15).until(
                EC.presence_of_element_located((By.XPATH, "//input[@type='email']"))
            )
            email_input.send_keys(email)
            email_input.send_keys(Keys.ENTER)
            time.sleep(10)

            # Enter password
            password_input = WebDriverWait(driver, 15).until(
                EC.presence_of_element_located((By.XPATH, "//input[@type='password']"))
            )
            password_input.send_keys(password)
            password_input.send_keys(Keys.ENTER)
            time.sleep(5)
            print("Logged in successfully!!!")
            # save_cookies(driver, cookies_path)
        except Exception as e:
            print(f"Error in feedly_login: {str(e)}")

def login_to_chatgpt_com(driver):
    try:
        login_button = WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.XPATH, '//button[@data-testid="login-button"]'))
        )
        login_button.click()
        time.sleep(5)
    except:
        print("Login button not found on chatgpt.com.")
        return
    # Attemp to login
    try:
        email = WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.XPATH, '//input[@id="email-input"]'))
        )#driver.find_element(By.XPATH, "//input[@id='email-input']")
        email.click()
        email.send_keys("rory@thebestreputation.com")
        email.send_keys(Keys.RETURN)
        time.sleep(5)
    except Exception as e:
        print("Unable to process username! Error: " + str(e))
        return
    # Attemp password
    try:
        passwd = WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.XPATH, '//input[@id="password"]'))
        )
        passwd.click()
        passwd.send_keys("Barbeque2045!")
        passwd.send_keys(Keys.RETURN)
        time.sleep(5)
    except Exception as e:
        print("Unable to process password to login! Error: " + str(e))
        return

def scroll_down(driver, element_selector):
    try:
        # Find the element (e.g., <main> tag or another container) and scroll within it
        main_element = WebDriverWait(driver, 5).until(
            EC.visibility_of_element_located((By.XPATH, element_selector))
        )
        # Scroll to the bottom of the element
        driver.execute_script(
            "arguments[0].scrollTo(0, arguments[0].scrollHeight);", main_element)
        time.sleep(1.5)
    except Exception as e:
        pass

def scrape_today_articles(driver):
    """Improved article scraping with reliable scrolling"""
    new_articles = set()
    scroll_attempts = 0
    max_scrolls = 30
    found_old_article = False
    
    try:
        for feed_url in [
            'https://feedly.com/i/collection/content/user/9e62dc2d-90e6-453b-88f4-47b630b9a4aa/category/global.all'
        ]:
            driver.get(feed_url)
            frame = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.ID, "feedlyFrame"))
            )
            
            last_height = driver.execute_script("return document.querySelector('#feedlyFrame').scrollHeight")
            
            while scroll_attempts < max_scrolls and not found_old_article:
                # Improved scrolling with explicit waits
                driver.execute_script("""
                    var frame = document.querySelector('#feedlyFrame');
                    frame.scrollTo({
                        top: frame.scrollHeight,
                        behavior: 'smooth'
                    });
                """)
                
                # Wait for content to load
                time.sleep(2)
                
                # Get current articles
                articles = driver.find_elements(By.CLASS_NAME, "entry.magazine")
                print(f"Found {len(articles)} articles on scroll {scroll_attempts + 1}")
                
                for article in articles:
                    try:
                        # Use WebDriverWait for reliable element access
                        date_span = WebDriverWait(article, 5).until(
                            EC.presence_of_element_located((By.CSS_SELECTOR, "span[title*='Published']"))
                        )
                        article_date = process_article_date(date_span.get_attribute("title"))
                        
                        if start_range <= article_date <= new_today_str:
                            title = article.find_element(By.CSS_SELECTOR, "a.EntryTitleLink").text
                            link = article.find_element(By.CSS_SELECTOR, "a.EntryTitleLink").get_attribute("href")
                            if (title, link) not in new_articles:
                                new_articles.add((title, link))
                                print(f"Found new article: {title}")
                        elif article_date < start_range:
                            found_old_article = True
                            print(f"Found article older than {start_range}")
                            break
                    except Exception as e:
                        continue

                # Check scroll progress
                new_height = driver.execute_script("return document.querySelector('#feedlyFrame').scrollHeight")
                if new_height == last_height:
                    # Wait longer and check again
                    time.sleep(3)
                    new_height = driver.execute_script("return document.querySelector('#feedlyFrame').scrollHeight")
                    if new_height == last_height:
                        print("No new content after waiting, checking dates...")
                        # Only break if we've found articles older than our target range
                        if found_old_article:
                            break
                
                last_height = new_height
                scroll_attempts += 1
                print(f"Completed scroll {scroll_attempts}/{max_scrolls}")

            logging.info(f"Completed {scroll_attempts} scrolls, found {len(new_articles)} articles")
            
    except Exception as e:
        logging.error(f"Scraping error: {e}")
        print(f"Error during scraping: {e}")
        
    return list(new_articles)

def get_article_content(url, timeout=10, max_retries=2):
    """Optimized article content fetching"""
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9",
    }
    
    session = requests.Session()  # Use session for connection pooling
    
    for attempt in range(max_retries):
        try:
            response = session.get(url, headers=headers, timeout=timeout)
            if response.status_code == 200:
                parser = HTMLParser(response.text)
                
                # Try all selectors at once
                selectors = "article, .article-content, .post-content, main, .entry-content"
                if content := parser.css_first(selectors):
                    return ' '.join(content.text().split())[:2000]
                return None
        except Exception as e:
            if attempt == max_retries - 1:
                print(f"Failed to fetch {url}: {e}")
            time.sleep(1)
    return None


def process_article_date(date_string):
    """Process article date string"""
    new_date = re.sub(r'\n', '', date_string).strip()
    new_date = re.sub(r'.*Received: | GMT.*', '', new_date).strip()
    if new_date[17:19] == '24':
        new_date = new_date[:17] + '00' + new_date[19:]
    return datetime.datetime.strptime(new_date, "%a, %d %b %Y %H:%M:%S")

def process_articles_batch(unique_new_articles, batch_size=50):
    """Optimized batch processing with progress updates"""
    decoded_articles = []
    titles = set()
    titles_neg = set()
    pg_links = []
    gpt_results = []
    
    print(f"\nStarting concurrent article processing for {len(unique_new_articles)} articles...")
    
    # Process articles concurrently
    with ThreadPoolExecutor(max_workers=5) as executor:
        future_to_article = {
            executor.submit(process_single_article, article): article 
            for article in unique_new_articles
        }
        
        completed = 0
        for future in as_completed(future_to_article):
            completed += 1
            try:
                result = future.result()
                if result:
                    if result.get('type') == 'negative':
                        titles_neg.add(result['title'])
                        print(f"Article marked negative ({completed}/{len(unique_new_articles)})")
                    else:
                        decoded_articles.append(result)
                        titles.add(result['title'])
                        pg_links.append(result['url'])
                        print(f"Article decoded successfully ({completed}/{len(unique_new_articles)})")
            except Exception as e:
                print(f"Error processing article: {str(e)}")

    print(f"\nStarting GPT processing for {len(decoded_articles)} articles...")
    
    # Process GPT in batches
    if decoded_articles:
        gpt_batch_size = 5
        for i in range(0, len(decoded_articles), gpt_batch_size):
            batch = decoded_articles[i:i + gpt_batch_size]
            try:
                print(f"\nProcessing GPT batch {i//gpt_batch_size + 1}/{(len(decoded_articles) + gpt_batch_size - 1)//gpt_batch_size}")
                batch_results = process_gpt_batch(batch)
                gpt_results.extend(batch_results)
                print(f"Successfully processed batch with {len(batch_results)} results")
            except Exception as e:
                print(f"GPT batch processing error: {e}")

    print(f"\nProcessing complete:")
    print(f"- Decoded articles: {len(decoded_articles)}")
    print(f"- GPT results: {len(gpt_results)}")
    print(f"- Negative articles: {len(titles_neg)}")

    return {
        'decoded_articles': decoded_articles,
        'titles': list(titles),
        'pg_links': pg_links,
        'titles_neg': list(titles_neg),
        'gpt_results': gpt_results
    }

def process_single_article(article):
    """Process individual article with content fetching"""
    title, url = str(article[0]), str(article[1])
    
    if title in titles_read or title in negative_titles_read:
        return None
        
    if is_check_title_against_keywords(title, negative_keywords):
        return {'type': 'negative', 'title': title}
        
    try:
        # Fix URL handling
        if 'news.google' in url:
            decoded = decode_google_news_url(url)
            final_url = decoded.get('decoded_url') if decoded.get('status') else url
        else:
            final_url = url
        
        print(f"Processing article: {final_url}")

        content = get_article_content(final_url)
        if content:
            return {
                'type': 'article',
                'title': title,
                'url': final_url,
                'content': content
            }
    except Exception as e:
        logging.error(f"Article processing error: {e}")
        
    return None

def process_gpt_batch(batch):
    """Process batch of articles with GPT-4 Turbo"""
    try:
        prompt = create_optimized_prompt(batch)
        client = OpenAI(api_key=OPEN_API_KEY)
        response = client.chat.completions.create(
            model="gpt-4-1106-preview",
            messages=[{"role": "user", "content": prompt}],
            max_tokens=1500,
            temperature=0.7,
            response_format={ "type": "text" }
        )
        
        if response.choices:
            result = parse_gpt_api_response_content(response.choices[0].message.content)
            # Add error checking for parsed result
            if result:
                try:
                    parsed_data = json.loads(result)
                    return parsed_data if isinstance(parsed_data, list) else []
                except json.JSONDecodeError:
                    print(f"Failed to parse GPT response: {result}")
                    return []
        print("No response from GPT")
        return []
        
    except Exception as e:
        print(f"GPT API error: {str(e)}\nPrompt: {prompt[:200]}...")
        return []
    
def create_optimized_prompt(batch):
    """Create optimized GPT prompt"""
    prompt = """Analyze these articles and provide structured summaries.
    Format each response exactly as: URL#TITLE#DESCRIPTION#REACH_OUT#REASONS#KEYWORDS#LOCATION

    Requirements:
    - Each response must be on a new line
    - Use exactly 6 '#' separators per line
    - Do not use '#' within field contents
    - Description: 100-200 word summary
    - Reach Out: Key person/organization
    - Reasons: Brief explanation
    - Keywords: 20 terms with |
    - Location: Specific place

    Example format:
    https://example.com#Article Title#Description text#Person Name#Reason text#keyword1|keyword2#City, State
    """
    
    for article in batch:
        prompt += f"\n\nArticle to analyze:\nURL: {article['url']}\nTitle: {article['title']}\nContent: {article['content'][:1000]}"
    
    return prompt

def main(email, password):
    initialize_global_variables()
    # Define headers and empty data
    if not os.path.exists(r"Rory Testing Sheet 2024.xlsx"):
        headers = ["Url","Title","Description","Reach Out","Reasons", "Keywords","Location","NOTES"]
        data = []
        df = pd.DataFrame(data, columns=headers)
        # Save to an Excel file
        df.to_excel(r"Rory Testing Sheet 2024.xlsx", index=False, engine="openpyxl")
        
        adjust_column_width(r"Rory Testing Sheet 2024.xlsx")
    test_offline = 0
    try:
        if not test_offline:
            feedly_login(driver, email, password)
            time.sleep(1.4) 
            articles = scrape_today_articles(driver)
        else:
            articles = []
            try:
                articles_raw_df = pd.read_csv(r'feedly_articles_raw.csv')
                articles = list(articles_raw_df.itertuples(index=False, name=None))
                print(f"Successfully read {len(articles)} articles from CSV")
            except Exception as e:
                print(f"Error reading from feedly_articles_raw.csv: {str(e)}")
                
        #   Save csv file for future offline testing
        save_for_offline_testing = 0
        if save_for_offline_testing:
            articles_raw_df = pd.DataFrame(articles, columns=['Title', 'URL'])
            articles_raw_df.to_csv(r'feedly_articles_raw.csv', index=False)
            print(f"Successfully wrote {len(articles)} articles to CSV")

        # Process articles
        if articles:
            print(f'Total collected articles: {len(articles)}')
            logging.info(f"'{len(articles)}' collected articles")

            # Convert to sets for faster lookup
            titles_read_set = set(titles_read)
            negative_titles_set = set(negative_titles_read)

            # First, filter out already processed titles
            unique_articles = [
                article for article in articles 
                if article[0] not in titles_read_set and article[0] not in negative_titles_set
            ]
            print(f'Unique articles (not in titles_read or negative_titles): {len(unique_articles)}')
            
            results = process_articles_batch(unique_articles)
            print(f"Processed {len(results['decoded_articles'])} articles")
            logging.info(f"'{len(results['decoded_articles'])}' processed articles")
            # Write to CSV immediately after processing
            try:
                articles_df = pd.DataFrame(results['decoded_articles'], columns=['Title', 'URL'])
                articles_df.to_csv(r'feedly_articles.csv', mode='a', index=False)
                print(f"Successfully wrote {len(results['decoded_articles'])} articles")
                
                # Create DataFrames for titles
                if results['titles']:
                    df_titles = pd.DataFrame({"Titles": results['titles']})
                    append_to_excel(r'titles_to_check.xlsx', df_titles, 'Sheet1')
                    
                if results['titles_neg']:
                    df_titles_neg = pd.DataFrame({"Titles": results['titles_neg']})
                    append_to_excel(r'negative_titles.xlsx', df_titles_neg, 'Sheet1')

                if results['gpt_results']:
                    # Map GPT results to Excel format
                    excel_data = pd.DataFrame([{
                        'URL': result['url'],
                        'Title': result['title'],
                        'Description': result['description'],
                        'Reach Out': result['reach_out'],
                        'Reasons': result['reasons'],
                        'Keywords': result['keywords'],
                        'Location': result['location'],
                        'NOTES': ''
                    } for result in results['gpt_results']])
                else:
                    excel_data = pd.DataFrame({
                        "URL": [article[1] for article in results['decoded_articles']],
                        "Title": [article[0] for article in results['decoded_articles']],
                        "Description": ["" for _ in results['decoded_articles']],
                        "Reach Out": ["" for _ in results['decoded_articles']],
                        "Reasons": ["" for _ in results['decoded_articles']],
                        "Keywords": ["" for _ in results['decoded_articles']],
                        "Location": ["" for _ in results['decoded_articles']],
                        "NOTES": ["" for _ in results['decoded_articles']]
                    })

                
                if not excel_data.empty:
                    append_to_excel(r"Rory Testing Sheet 2024.xlsx", excel_data, 'Sheet1')
                    print(f"Successfully wrote {len(excel_data)} articles to Excel")
                    logging.info(f"'{len(excel_data)}' articles to Excel\n")
                    
            except Exception as e:
                print(f"Error writing to files: {str(e)}")

    finally:
        if not test_offline:
            driver.execute_script("window.localStorage.clear();")
            driver.execute_script("window.sessionStorage.clear();")
            driver.execute_script("console.clear();")
            driver.quit()

if __name__ == "__main__":
    start = time.time()
    email = os.getenv("EMAIL")
    password = os.getenv("PASSWORD")
    main(email, password)
    end = time.time()
    print("Time Taken: {:.6f}s".format(end - start))
    time.sleep(30) # Delay the cmd window exiting immediately after a scheduled run
